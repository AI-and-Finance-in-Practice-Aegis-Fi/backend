"""
Seed script for Aegis-Fi database.
Reads 1차_데이터셋.xlsx and inserts all data.

Usage: python seed.py
"""
import asyncio
import os
import re
import warnings
from datetime import date, datetime, time, timedelta

import asyncpg
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# Excel stores dates as days since 1899-12-30 (with the 1900 leap-year bug)
_EXCEL_EPOCH = datetime(1899, 12, 30)

# Cells where openpyxl fails to parse huge date-serials (e.g. 3,297,000 KRW).
# Populated from openpyxl warnings when the workbook is loaded.
_cell_overrides: dict[str, int] = {}


def _load_workbook(path: str) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _strip_id(val: str) -> int:
    """'E001' → 1, 'U0023' → 23, 'T0500' → 500"""
    return int(re.sub(r"[A-Za-z]", "", str(val)))


def _to_int(val) -> int | None:
    """
    Excel stores large numbers in date-formatted cells, so openpyxl returns
    datetime objects for monetary values. Convert back to the actual integer.
    - datetime.time(0, 0)  → 0  (cell value was literally zero)
    - '#VALUE!' string     → caller must resolve via _cell_overrides
    """
    if val is None:
        return None
    if isinstance(val, str):
        return None  # error cell; caller resolves from _cell_overrides
    if isinstance(val, datetime):
        return int((val - _EXCEL_EPOCH).days)
    if isinstance(val, time):
        return 0
    return int(val)


def _to_date(val) -> date | None:
    """Excel integer serial or datetime → Python date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        return (_EXCEL_EPOCH + timedelta(days=int(val))).date()
    return None


def _to_datetime(val) -> datetime | None:
    """Excel float serial (days + fraction) → Python datetime with time."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return _EXCEL_EPOCH + timedelta(days=float(val))
    return None


def _load_sheet(wb, sheet_name: str) -> list[tuple]:
    """Read a sheet and capture openpyxl out-of-range date warnings into _cell_overrides."""
    ws = wb[sheet_name]
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    for w in caught:
        msg = str(w.message)
        m = re.search(r"Cell (\S+) is marked as a date but the serial value (\d+)", msg)
        if m:
            _cell_overrides[m.group(1)] = int(m.group(2))
    return [r for r in rows if any(v is not None for v in r)]


async def seed():
    url = os.getenv("DATABASE_URL", "")
    # asyncpg.connect uses postgresql:// scheme
    dsn = url.replace("postgresql+asyncpg://", "postgresql://")

    print(f"Connecting to {dsn.split('@')[-1]} ...")
    conn = await asyncpg.connect(dsn=dsn)

    wb = _load_workbook("1차_데이터셋.xlsx")

    try:
        async with conn.transaction():

            # ── 1. department ──────────────────────────────────────────────
            # Sheet: Department_Budget
            # Columns: department_id, department_name, monthly_limit,
            #          current_spend, remaining_budget, spend_rate
            rows = _load_sheet(wb, "Department_Budget")
            for r in rows:
                await conn.execute(
                    """
                    INSERT INTO department
                        (department_id, department_name, monthly_budget_limit, current_spending)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (department_id) DO NOTHING
                    """,
                    _strip_id(r[0]), r[1], r[2], r[3],
                )
            print(f"  department:        {len(rows):>4} rows")

            # ── 2. employee ────────────────────────────────────────────────
            # Sheet: Employee
            # Columns: employee_id, employee_name, department_id,
            #          department_name, position, hire_date, employment_type
            rows = _load_sheet(wb, "Employee")
            for r in rows:
                await conn.execute(
                    """
                    INSERT INTO employee
                        (employee_id, employee_name, position, department_id)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (employee_id) DO NOTHING
                    """,
                    _strip_id(r[0]), r[1], r[4], _strip_id(r[2]),
                )
            print(f"  employee:          {len(rows):>4} rows")

            # ── 3. saas_subscription ───────────────────────────────────────
            # Sheet: SaaS_Subscription
            # Columns: subscription_id, department_id, department_name,
            #          product_name, vendor, monthly_fee, seat_count,
            #          active_users, waste_amount, risk_level, renewal_date
            rows = _load_sheet(wb, "SaaS_Subscription")
            for r in rows:
                risk = (r[9] or "LOW").strip().upper()
                await conn.execute(
                    """
                    INSERT INTO saas_subscription
                        (subscription_id, subscription_name, provider,
                         monthly_fee, total_seats, used_seats,
                         wasted_amount, risk_level, renewal_date, department_id)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8::risk_level_enum, $9, $10)
                    ON CONFLICT (subscription_id) DO NOTHING
                    """,
                    _strip_id(r[0]), r[3], r[4],
                    _to_int(r[5]), int(r[6]), int(r[7]),
                    _to_int(r[8]) or 0, risk, _to_date(r[10]), _strip_id(r[1]),
                )
            print(f"  saas_subscription: {len(rows):>4} rows")

            # ── 4. saas_usage ──────────────────────────────────────────────
            # Sheet: SaaS_Usage
            # Columns: usage_id, employee_id, employee_name, department_id,
            #          department_name, position, subscription_id, saas_name,
            #          last_login_date, monthly_usage_count, is_ghost_account
            rows = _load_sheet(wb, "SaaS_Usage")
            for r in rows:
                await conn.execute(
                    """
                    INSERT INTO saas_usage
                        (usage_id, employee_id, subscription_id,
                         last_login_date, monthly_usage_count, is_ghost_account)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT (usage_id) DO NOTHING
                    """,
                    _strip_id(r[0]), _strip_id(r[1]), _strip_id(r[6]),
                    _to_date(r[8]), int(r[9]) if r[9] is not None else 0,
                    bool(r[10]) if r[10] is not None else False,
                )
            print(f"  saas_usage:        {len(rows):>4} rows")

            # ── 5. transaction ─────────────────────────────────────────────
            # Sheet: Transactions
            # Columns: transaction_id, employee_id, employee_name,
            #          department_id, department_name, merchant_name,
            #          category, amount, purpose_text, approved,
            #          ai_risk_score, ai_reason, tx_timestamp
            rows = _load_sheet(wb, "Transactions")
            # Map sheet row number → amount for cells openpyxl returned '#VALUE!'.
            # Sheet row 273 = rows[271] because data starts at sheet row 2 (index 0).
            _tx_amount_overrides = {
                int(coord[1:]) - 2: val   # "H273" → rows[271]
                for coord, val in _cell_overrides.items()
                if coord.startswith("H")
            }
            if _tx_amount_overrides:
                print(f"  [warn] restoring {len(_tx_amount_overrides)} out-of-range amount(s): {_tx_amount_overrides}")
            for idx, r in enumerate(rows):
                amount = _to_int(r[7])
                if amount is None:
                    amount = _tx_amount_overrides.get(idx)
                await conn.execute(
                    """
                    INSERT INTO transaction
                        (transaction_id, employee_id, merchant_name, category,
                         amount, user_input_reason, is_approved,
                         ai_risk_score, ai_risk_reason, payment_time)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                    ON CONFLICT (transaction_id) DO NOTHING
                    """,
                    _strip_id(r[0]), _strip_id(r[1]), r[5], r[6],
                    amount, r[8],
                    bool(r[9]) if r[9] is not None else None,
                    float(r[10]) if r[10] is not None else None,
                    r[11], _to_datetime(r[12]),
                )
            print(f"  transaction:       {len(rows):>4} rows")

            # ── 6. ai_payment_policy ───────────────────────────────────────
            # Sheet: AI_Payment_Policy
            # Columns: policy_id, department_id, department_name,
            #          restricted_category, single_payment_limit, is_blocked
            rows = _load_sheet(wb, "AI_Payment_Policy")
            for r in rows:
                await conn.execute(
                    """
                    INSERT INTO ai_payment_policy
                        (policy_id, employee_id, department_id,
                         restricted_category, single_payment_limit, is_blocked)
                    VALUES ($1, NULL, $2, $3, $4, $5)
                    ON CONFLICT (policy_id) DO NOTHING
                    """,
                    _strip_id(r[0]), _strip_id(r[1]),
                    r[3], _to_int(r[4]),
                    bool(r[5]) if r[5] is not None else False,
                )
            print(f"  ai_payment_policy: {len(rows):>4} rows")

            # ── 7. approval_log ────────────────────────────────────────────
            # Sheet: Approvals
            # Columns: approval_id, transaction_id, handler_employee_id,
            #          handler_employee_name, decision, decision_time, handler_note
            rows = _load_sheet(wb, "Approvals")
            for r in rows:
                await conn.execute(
                    """
                    INSERT INTO approval_log
                        (approval_id, transaction_id, approver_employee_id,
                         approval_result, approval_time, approval_reason)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT (approval_id) DO NOTHING
                    """,
                    _strip_id(r[0]), _strip_id(r[1]), _strip_id(r[2]),
                    r[4] == "승인", _to_datetime(r[5]), r[6],
                )
            print(f"  approval_log:      {len(rows):>4} rows")

        print("\nSeed completed successfully.")

        # ── Row count verification ─────────────────────────────────────────
        print("\nRow counts:")
        for table in [
            "department", "employee", "saas_subscription",
            "saas_usage", "transaction", "ai_payment_policy", "approval_log",
        ]:
            n = await conn.fetchval(f"SELECT COUNT(*) FROM {table}")
            print(f"  {table:<22} {n:>4}")

    finally:
        wb.close()
        await conn.close()


if __name__ == "__main__":
    asyncio.run(seed())
